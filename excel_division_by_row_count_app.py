from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from copy import copy
import tkinter as tk
from tkinter import filedialog
import requests
from io import BytesIO
from PIL import Image, ImageTk
import math

root = tk.Tk()
root.title("Excel Division by row count")

# Create and place labels and entries
tk.Label(root, text="Path to excel:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
open_name_var = tk.StringVar()
tk.Entry(root, textvariable=open_name_var, width=50).grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Row count (e.g. 2 results in excels with 2 rows (+header)):").grid(row=1, column=0, padx=10, pady=5, sticky="e")
row_count_var = tk.IntVar()
tk.Entry(root, textvariable=row_count_var, width=50).grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Name of the worksheet:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
sheet_name_var = tk.StringVar()
tk.Entry(root, textvariable=sheet_name_var, width=50).grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Name of the new worksheet:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
sheet_name_new_var = tk.StringVar()
tk.Entry(root, textvariable=sheet_name_new_var, width=50).grid(row=3, column=1, padx=10, pady=5)

tk.Label(root, text="Folder to save the divided excels:").grid(row=4, column=0, padx=10, pady=5, sticky="e")
save_name_var = tk.StringVar()
tk.Entry(root, textvariable=save_name_var, width=50).grid(row=4, column=1, padx=10, pady=5)

tk.Label(root, text="Put in the name of the new excels:").grid(row=5, column=0, padx=10, pady=5, sticky="e")
excel_name_var = tk.StringVar()
tk.Entry(root, textvariable=excel_name_var, width=50).grid(row=5, column=1, padx=10, pady=5)

# Button to open file dialog
tk.Button(root, text="Browse", command=lambda: open_name_var.set(filedialog.askopenfilename())).grid(row=0, column=2, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: save_name_var.set(filedialog.askdirectory())).grid(row=4, column=2, padx=10, pady=5)

try:
    response = requests.get("https://raw.githubusercontent.com/MMateo1120/Excel_division_app/refs/heads/main/mz6axvogusy31.jpg")
    image_data = BytesIO(response.content)
    image = Image.open(image_data)
    resized_image = image.resize((200, int(image.height * 200 / image.width)))
    cat_image = ImageTk.PhotoImage(resized_image)
    tk.Label(root, image=cat_image).grid(row=0, column=6, rowspan=5, sticky="nwe")
except:
    pass

# Function to submit and close
def submit():
    global open_name, sheet_name, sheet_name_new, row_count, save_name, excel_name
    open_name = open_name_var.get()
    row_count = row_count_var.get()
    sheet_name = sheet_name_var.get()
    sheet_name_new = sheet_name_new_var.get()
    save_name = save_name_var.get()
    excel_name = excel_name_var.get()
    root.destroy()

# Submit button
tk.Button(root, text="Submit", command=submit).grid(row=6, column=2, padx=10, pady=10)
root.mainloop()

#Excel load
wb = load_workbook(open_name)
ws = wb[sheet_name]

#Get the header
header = []
for cell_row in ws[1]:
    header.append(cell_row.value)

#row_count
div_num = math.ceil(ws.max_row/row_count)
mins = [i*row_count+2 for i in range(div_num)]
maxs = [i*row_count+row_count+1 for i in range(div_num)]
maxs[-1] = ws.max_row

for i,j in zip(mins,maxs):
    wb_new = Workbook()
    ws_new = wb_new.create_sheet(sheet_name_new)
    del wb_new["Sheet"]
    
    ws_new.append(header)
    for cell_orig, cell_new in zip(ws[1], ws_new[1]):
        cell_new.alignment = copy(cell_orig.alignment)
        cell_new.number_format = copy(cell_orig.number_format)
        cell_new.font = copy(cell_orig.font)
        cell_new.fill = copy(cell_orig.fill)
        cell_new.border = copy(cell_orig.border)
        
    for col_num in range(1,ws.max_column+1):
        ws_new.column_dimensions[get_column_letter(col_num)].width = ws.column_dimensions[get_column_letter(col_num)].width
        
    # Iterate over the rows in the worksheet
    for row in ws.iter_rows(min_row=i, max_row=j, values_only=False):
        print(f"{row[0].row}.row from the total of {ws.max_row} rows")
 
        # Append the row to the workbook
        ws_new.append([cell.value for cell in row])

        # Copy styles
        for cell_orig, cell_new in zip(row, ws_new[ws_new.max_row]):
            cell_new.alignment = copy(cell_orig.alignment)
            cell_new.number_format = copy(cell_orig.number_format)
            cell_new.font = copy(cell_orig.font)
            cell_new.fill = copy(cell_orig.fill)
            cell_new.border = copy(cell_orig.border)

    wb_new.save(save_name + f"/{i}_{j}_{excel_name}.xlsx")


